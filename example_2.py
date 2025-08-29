import csv
from pathlib import Path

from faker import Faker
from faker.providers import internet
from faker.providers.address.ru_RU import Provider as RuAddress
from faker.providers.person.ru_RU import Provider as RuPerson
from openpyxl import load_workbook

XLSX_FILE_PATH = Path("files") / 'qa.xlsx'


# wb = load_workbook(XLSX_FILE_PATH)
# ws = wb.active
#
# header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
#
# for row in ws.iter_rows(min_row=2, values_only=True):
#     print(dict(zip(header, row)))
#
fake = Faker()
fake.add_provider(internet)
fake.add_provider(RuAddress)
fake.add_provider(RuPerson)

CSV_FILE_PATH = Path("files") / 'users.csv'
with open(CSV_FILE_PATH, 'w', newline='') as csv_file:
    writer = csv.writer(csv_file, delimiter=',', )

    writer.writerow(['id', 'name', 'address', 'ip'])
    for number in range(10):
        writer.writerow([str(number).zfill(5), fake.name(), fake.address(), fake.ipv4_public()])
